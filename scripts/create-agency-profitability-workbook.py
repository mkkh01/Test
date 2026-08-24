from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook
from pathlib import Path

out = Path('/home/ubuntu/Test/product-assets/client-payment-scope-kit/12_AGENCY_PROFITABILITY_WORKSHEET.xlsx')
wb = Workbook()
ws = wb.active
ws.title = 'Instructions'

navy = '17324D'
blue = '2F75B5'
light_blue = 'D9EAF7'
green = 'E2F0D9'
orange = 'FCE4D6'
grey = 'F2F2F2'
white = 'FFFFFF'
thin_grey = Side(style='thin', color='D9E1F2')

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

ws['A1'] = 'Agency Profitability Worksheet'
ws['A1'].font = Font(bold=True, size=18, color=white)
ws['A1'].fill = PatternFill('solid', fgColor=navy)
ws.merge_cells('A1:F1')
ws['A3'] = 'Use this workbook to forecast margin before accepting work and to compare the approved plan with actual delivery.'
ws['A3'].alignment = Alignment(wrap_text=True)
ws.merge_cells('A3:F3')
ws['A5'] = 'Workflow'
ws['A5'].font = Font(bold=True, color=white)
ws['A5'].fill = PatternFill('solid', fgColor=blue)
ws.merge_cells('A5:F5')
workflow = [
    ('1', 'Add one row per project on Project P&L.'),
    ('2', 'Enter revenue, planned hours, hourly internal cost, and other direct costs.'),
    ('3', 'Update actual hours and costs weekly; formulas calculate variance and margin.'),
    ('4', 'Use Capacity Plan before accepting new work.'),
    ('5', 'Use Change Impact to price extra work before production begins.'),
]
for row, (step, text) in enumerate(workflow, 6):
    ws.cell(row, 1, step).font = Font(bold=True, color=blue)
    ws.cell(row, 2, text)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
ws['A13'] = 'Definitions'
ws['A13'].font = Font(bold=True, color=white)
ws['A13'].fill = PatternFill('solid', fgColor=blue)
ws.merge_cells('A13:F13')
defs = [
    ('Gross margin %', '(Revenue − direct cost) ÷ Revenue.'),
    ('Forecast margin %', '(Revenue − forecast direct cost) ÷ Revenue.'),
    ('Hours variance', 'Actual hours − planned hours. Positive means over plan.'),
    ('Change request', 'Any added deliverable, revision, channel, meeting, or deadline change outside the approved baseline.'),
]
for row, (name, definition) in enumerate(defs, 14):
    ws.cell(row, 1, name).font = Font(bold=True)
    ws.cell(row, 2, definition)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
for col, width in {'A':22, 'B':48, 'C':18, 'D':18, 'E':18, 'F':18}.items():
    ws.column_dimensions[col].width = width

pnl = wb.create_sheet('Project P&L')
pnl.sheet_view.showGridLines = False
headers = ['Project ID', 'Client', 'Service', 'Status', 'Revenue', 'Planned hours', 'Actual hours', 'Internal hourly cost', 'Other direct cost', 'Forecast cost', 'Actual cost', 'Forecast margin %', 'Actual margin %', 'Hours variance', 'Margin flag', 'Next action']
for col, header in enumerate(headers, 1):
    cell = pnl.cell(1, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=navy)
    cell.alignment = Alignment(wrap_text=True, horizontal='center')

for r in range(2, 102):
    pnl.cell(r, 10, f'=IF(E{r}="","",F{r}*H{r}+I{r})')
    pnl.cell(r, 11, f'=IF(E{r}="","",G{r}*H{r}+I{r})')
    pnl.cell(r, 12, f'=IFERROR((E{r}-J{r})/E{r},"")')
    pnl.cell(r, 13, f'=IFERROR((E{r}-K{r})/E{r},"")')
    pnl.cell(r, 14, f'=IF(E{r}="","",G{r}-F{r})')
    pnl.cell(r, 15, f'=IF(E{r}="","",IF(M{r}<0.45,"RED",IF(M{r}<0.6,"AMBER","GREEN")))')
    for c in range(1, 17):
        pnl.cell(r, c).border = Border(bottom=thin_grey)
        pnl.cell(r, c).alignment = Alignment(vertical='top', wrap_text=True)
    for c in [5, 8, 9, 10, 11]:
        pnl.cell(r, c).number_format = '0.00'
    for c in [12, 13]:
        pnl.cell(r, c).number_format = '0.0%'

pnl.freeze_panes = 'A2'
pnl.auto_filter.ref = 'A1:P101'
for c, width in {1:12,2:20,3:22,4:16,5:13,6:14,7:13,8:18,9:16,10:15,11:13,12:15,13:13,14:14,15:12,16:28}.items():
    pnl.column_dimensions[get_column_letter(c)].width = width
status_dv = DataValidation(type='list', formula1='"Pipeline,Active,Review,Handover,Closed,Paused"', allow_blank=True)
pnl.add_data_validation(status_dv)
status_dv.add('D2:D101')
pnl.conditional_formatting.add('O2:O101', FormulaRule(formula=['O2="RED"'], fill=PatternFill('solid', fgColor='F4CCCC')))
pnl.conditional_formatting.add('O2:O101', FormulaRule(formula=['O2="AMBER"'], fill=PatternFill('solid', fgColor='FCE5CD')))
pnl.conditional_formatting.add('O2:O101', FormulaRule(formula=['O2="GREEN"'], fill=PatternFill('solid', fgColor='D9EAD3')))

cap = wb.create_sheet('Capacity Plan')
cap.sheet_view.showGridLines = False
cap_headers = ['Team member', 'Role', 'Weekly capacity hours', 'Utilization target %', 'Committed hours', 'Reserved ops hours', 'Available delivery hours', 'Overallocated?']
for col, header in enumerate(cap_headers, 1):
    cell = cap.cell(1, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=navy)
    cell.alignment = Alignment(wrap_text=True, horizontal='center')
for r in range(2, 22):
    cap.cell(r, 7, f'=IF(A{r}="","",C{r}*D{r}-E{r}-F{r})')
    cap.cell(r, 8, f'=IF(A{r}="","",IF(G{r}<0,"YES","NO"))')
    cap.cell(r, 4).number_format = '0%'
    for c in range(1, 9):
        cap.cell(r, c).border = Border(bottom=thin_grey)
        cap.cell(r, c).alignment = Alignment(wrap_text=True)
cap.freeze_panes = 'A2'
for c, width in {1:20,2:20,3:20,4:20,5:18,6:18,7:22,8:16}.items():
    cap.column_dimensions[get_column_letter(c)].width = width
cap.conditional_formatting.add('H2:H21', FormulaRule(formula=['H2="YES"'], fill=PatternFill('solid', fgColor='F4CCCC')))

change = wb.create_sheet('Change Impact')
change.sheet_view.showGridLines = False
change_headers = ['Change ID', 'Project ID', 'Requested item', 'Extra hours', 'Hourly sell rate', 'Added price', 'Internal cost', 'Net contribution', 'Approval status', 'Decision date']
for col, header in enumerate(change_headers, 1):
    cell = change.cell(1, col, header)
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=navy)
    cell.alignment = Alignment(wrap_text=True, horizontal='center')
for r in range(2, 52):
    change.cell(r, 6, f'=IF(D{r}="","",D{r}*E{r})')
    change.cell(r, 7, f'=IF(D{r}="","",D{r}*0)')
    change.cell(r, 8, f'=IF(F{r}="","",F{r}-G{r})')
    for c in range(1, 11):
        change.cell(r, c).border = Border(bottom=thin_grey)
        change.cell(r, c).alignment = Alignment(wrap_text=True)
change.freeze_panes = 'A2'
for c, width in {1:12,2:14,3:34,4:14,5:18,6:14,7:14,8:18,9:18,10:16}.items():
    change.column_dimensions[get_column_letter(c)].width = width
approval_dv = DataValidation(type='list', formula1='"Draft,Awaiting approval,Approved,Rejected,Invoiced"', allow_blank=True)
change.add_data_validation(approval_dv)
approval_dv.add('I2:I51')

summary = wb.create_sheet('Dashboard')
summary.sheet_view.showGridLines = False
summary['A1'] = 'Agency Profitability Dashboard'
summary['A1'].font = Font(bold=True, size=18, color=white)
summary['A1'].fill = PatternFill('solid', fgColor=navy)
summary.merge_cells('A1:F1')
metrics = [
    ('Total revenue', '=SUM(\'Project P&L\'!E2:E101)', '0.00'),
    ('Forecast direct cost', '=SUM(\'Project P&L\'!J2:J101)', '0.00'),
    ('Actual direct cost', '=SUM(\'Project P&L\'!K2:K101)', '0.00'),
    ('Forecast margin %', '=IFERROR((B3-B4)/B3,"")', '0.0%'),
    ('Actual margin %', '=IFERROR((B3-B5)/B3,"")', '0.0%'),
    ('Planned hours', '=SUM(\'Project P&L\'!F2:F101)', '0.00'),
    ('Actual hours', '=SUM(\'Project P&L\'!G2:G101)', '0.00'),
    ('Open change value', '=SUMIF(\'Change Impact\'!I2:I51,"<>Approved",\'Change Impact\'!F2:F51)', '0.00'),
]
for idx, (label, formula, fmt) in enumerate(metrics, 3):
    summary.cell(idx, 1, label).font = Font(bold=True)
    summary.cell(idx, 2, formula).number_format = fmt
    summary.cell(idx, 1).fill = PatternFill('solid', fgColor=light_blue)
    summary.cell(idx, 2).fill = PatternFill('solid', fgColor=grey)
    summary.cell(idx, 1).border = Border(bottom=thin_grey)
    summary.cell(idx, 2).border = Border(bottom=thin_grey)
summary['D3'] = 'Project'
summary['E3'] = 'Actual margin %'
for cell in summary[3][3:5]:
    cell.font = Font(bold=True, color=white)
    cell.fill = PatternFill('solid', fgColor=blue)
for r in range(4, 14):
    source_r = r - 2
    summary.cell(r, 4, f'=IF(\'Project P&L\'!A{source_r}="","",\'Project P&L\'!A{source_r})')
    summary.cell(r, 5, f'=IF(\'Project P&L\'!A{source_r}="","",\'Project P&L\'!M{source_r})')
    summary.cell(r, 5).number_format = '0.0%'
chart = BarChart()
chart.title = 'Actual margin by project'
chart.y_axis.title = 'Margin %'
chart.x_axis.title = 'Project'
data = Reference(summary, min_col=5, min_row=3, max_row=13)
cats = Reference(summary, min_col=4, min_row=4, max_row=13)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 13
summary.add_chart(chart, 'D15')
for col, width in {'A':25, 'B':18, 'C':4, 'D':18, 'E':18, 'F':18}.items():
    summary.column_dimensions[col].width = width

for sheet in wb.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and cell.alignment == Alignment():
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

wb.save(out)
print(out)
